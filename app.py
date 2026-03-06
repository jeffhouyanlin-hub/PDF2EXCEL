from __future__ import annotations

import base64
import io
import json
import re
import tempfile
import zipfile
from pathlib import Path

import requests

import pandas as pd
import streamlit as st
import streamlit.components.v1 as stc
from openpyxl.utils import get_column_letter

from core.batch import BatchProcessor
from core.converter import ExcelConverter
from core.extractor import PDFExtractor
from core.statement_parser import BankStatementParser, RowPosition
from core.text_extractor import TextBasedExtractor
from core.verifier import (
    ArithmeticIssue,
    DataVerifier,
    LayerResult,
    TripleVerificationResult,
    TripleVerifier,
    _normalize,
)

st.set_page_config(page_title="PDF2EXCEL", page_icon="📊", layout="wide")

# --- 密码验证 / Password Auth ---
def _check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True
    pwd = st.text_input("请输入访问密码 / Enter password", type="password", key="pwd_input")
    if not pwd:
        st.stop()
    if pwd == st.secrets.get("APP_PASSWORD", ""):
        st.session_state.authenticated = True
        st.rerun()
    else:
        st.error("密码错误 / Incorrect password")
        st.stop()
    return False

_check_password()

st.title("📊 PDF2EXCEL")
st.caption("PDF 表格 → Excel 批量转换工具 / PDF Table → Excel Batch Converter")

# --- 侧边栏 / Sidebar ---
_MODE_LABELS = {
    "自动检测": "自动检测 / Auto Detect",
    "标准表格": "标准表格 / Standard Table",
    "银行账单": "银行账单 / Bank Statement",
}
_SHEET_LABELS = {
    "每个表格一个 Sheet": "每个表格一个 Sheet / One Sheet per Table",
    "合并到一个 Sheet": "合并到一个 Sheet / Merge into One Sheet",
}

with st.sidebar:
    st.header("设置 / Settings")
    parse_mode = st.radio(
        "解析模式 / Parse Mode",
        list(_MODE_LABELS.keys()),
        index=0,
        format_func=lambda x: _MODE_LABELS[x],
        help="自动检测：先尝试标准表格提取，失败则用银行账单解析器 / Auto: try standard first, fallback to bank statement",
    )
    sheet_mode = st.radio(
        "Sheet 策略 / Sheet Strategy",
        list(_SHEET_LABELS.keys()),
        index=0,
        format_func=lambda x: _SHEET_LABELS[x],
    )
    sheet_per_table = sheet_mode == "每个表格一个 Sheet"
    max_workers = st.slider("并行线程数 / Parallel Threads", 1, 15, 4)
    st.divider()
    st.caption("© Dr. Jeff Hou · v0.4.4")
    if st.button("📮 报错反馈 / Error Feedback", use_container_width=True):
        st.session_state.show_feedback = True
    if st.button("🔍 数据复核 / Data Verification", use_container_width=True):
        st.session_state.show_verification = True

# --- 意见反馈页面 / Feedback Page ---
if st.session_state.get("show_feedback"):
    st.subheader("📮 意见反馈 / Error Feedback")
    st.info(
        "若您的对账单无法识别或识别错误，请提交样本和错误描述，我们会安排优化解决。\n\n"
        "If your statement cannot be recognized or is incorrectly parsed, "
        "please submit a sample and error description. We will arrange optimization."
    )

    with st.form("feedback_form", clear_on_submit=True):
        feedback_file = st.file_uploader(
            "上传 PDF 样本 / Upload PDF Sample", type=["pdf"], key="feedback_pdf"
        )
        feedback_text = st.text_area(
            "错误描述 / Error Description",
            placeholder="请描述您遇到的问题... / Please describe the issue...",
        )
        col_submit, col_back = st.columns(2)
        with col_submit:
            submitted = st.form_submit_button(
                "提交反馈 / Submit", type="primary", use_container_width=True
            )
        with col_back:
            back = st.form_submit_button("返回 / Back", use_container_width=True)

    if back:
        st.session_state.show_feedback = False
        st.rerun()

    if submitted:
        if not feedback_file:
            st.warning("请上传 PDF 文件。/ Please upload a PDF file.")
        elif not feedback_text.strip():
            st.warning("请填写错误描述。/ Please fill in the error description.")
        else:
            formspree_id = st.secrets.get("FORMSPREE_ID", "")
            if not formspree_id:
                st.error("反馈服务未配置，请联系管理员。/ Feedback service not configured.")
            else:
                import base64
                with st.spinner("正在提交... / Submitting..."):
                    pdf_b64 = base64.b64encode(feedback_file.getvalue()).decode()
                    resp = requests.post(
                        f"https://formspree.io/f/{formspree_id}",
                        json={
                            "email": "feedback@pdf2excel.app",
                            "message": feedback_text,
                            "_subject": f"PDF2EXCEL 反馈: {feedback_file.name}",
                            "filename": feedback_file.name,
                            "filesize": f"{len(feedback_file.getvalue()) / 1024:.1f} KB",
                            "pdf_data": pdf_b64,
                        },
                        headers={"Accept": "application/json"},
                    )
                if resp.ok:
                    st.success("反馈已提交，感谢您的支持！/ Feedback submitted, thank you!")
                else:
                    st.error(
                        f"提交失败（{resp.status_code}），请稍后重试。/ "
                        f"Failed ({resp.status_code}), please try again."
                    )

    st.stop()


def _build_verification_payload(
    unified_rows: list[dict],
    row_positions: list,
    row_headers: list[str],
    is_statement: bool,
    summary_positions: list | None = None,
    layers: list[LayerResult] | None = None,
) -> dict:
    """Build data payload for the dual-panel verification component.

    Returns a dict with segments, rows, comparisons, and PDF positions,
    ready to be serialized to JSON and embedded in the HTML component.
    """
    # --- Segments: group consecutive rows by sheet ---
    segments: list[dict] = []
    prev_sheet = None
    for i, row in enumerate(unified_rows):
        if row["sheet"] != prev_sheet:
            if segments:
                segments[-1]["endIdx"] = i
            segments.append({
                "sheet": row["sheet"],
                "columns": row["columns"],
                "startIdx": i,
            })
            prev_sheet = row["sheet"]
    if segments:
        segments[-1]["endIdx"] = len(unified_rows)

    # --- Rows: compact per-row data ---
    rows = []
    for i, ur in enumerate(unified_rows):
        seg_idx = next(
            j for j, s in enumerate(segments)
            if s["startIdx"] <= i < s["endIdx"]
        )
        cols = ur["columns"]
        rows.append({
            "seg": seg_idx,
            "ev": [ur["excel_vals"].get(c, "") for c in cols],
            "pv": [ur["pdf_vals"].get(c, "") for c in cols],
            "ok": len(ur["diff_cols"]) == 0,
            "dc": [ci for ci, c in enumerate(cols) if c in ur["diff_cols"]],
        })

    # --- PDF positions ---
    positions = []
    if is_statement and row_positions:
        for rp in row_positions:
            positions.append({
                "pg": rp.page_index,
                "yt": round(rp.y_top, 1),
                "yb": round(rp.y_bottom, 1),
                "ph": round(rp.page_height, 1),
            })

    # --- Summary PDF positions ---
    sum_positions = []
    if summary_positions:
        for sp in summary_positions:
            if sp is not None:
                sum_positions.append({
                    "pg": sp.page_index,
                    "yt": round(sp.y_top, 1),
                    "yb": round(sp.y_bottom, 1),
                    "ph": round(sp.page_height, 1),
                })
            else:
                sum_positions.append(None)

    # Build layers info for triple verification status bar
    layers_data = []
    if layers:
        for layer in layers:
            ld: dict = {
                "name": layer.name,
                "label": layer.label,
                "passed": layer.passed,
                "skipped": layer.skipped,
                "total_cells": layer.total_cells,
                "mismatched": layer.mismatched_cells,
            }
            if layer.arithmetic_issues:
                ld["arithmetic"] = [
                    {"check": ai.check, "row": ai.row,
                     "expected": ai.expected, "actual": ai.actual,
                     "message": ai.message}
                    for ai in layer.arithmetic_issues
                ]
            layers_data.append(ld)

    return {
        "segments": segments,
        "rows": rows,
        "pdfPos": positions,
        "summaryPdfPos": sum_positions,
        "pdfHeaders": row_headers,
        "isStmt": is_statement,
        "total": len(unified_rows),
        "matched": sum(1 for r in rows if r["ok"]),
        "mismatched": sum(1 for r in rows if not r["ok"]),
        "layers": layers_data,
    }


def _build_verification_html(payload_json: str, pdf_b64: str) -> str:
    """Build HTML for the dual-panel verification component."""
    tpl = (Path(__file__).parent / "templates" / "verification.html").read_text(encoding="utf-8")
    return tpl.replace('"__PAYLOAD__"', payload_json).replace("__PDF_B64__", pdf_b64)


def _build_merged_verification_payload(
    merged_excel_bytes: bytes,
    row_mapping: list[tuple[str, int]],
    verify_files: dict,
) -> dict:
    """Build payload for the merged verification component.

    Handles three row types:
    - separator rows (stem="__separator__"): empty rows between blocks
    - Opening Balance rows (orig_idx=-1): first row of each block
    - transaction rows: mapped back to source PDF for comparison
    """
    xls = pd.ExcelFile(io.BytesIO(merged_excel_bytes))
    edf = pd.read_excel(xls, sheet_name="Transactions")
    cols = list(edf.columns)

    segments = [{"sheet": "Transactions", "columns": cols, "startIdx": 0, "endIdx": len(edf)}]
    rows = []
    tx_total = 0  # count only actual transaction rows for stats

    for i in range(len(edf)):
        ev = {}
        for c in cols:
            raw = str(edf.iloc[i][c]) if str(edf.iloc[i][c]) != "nan" else ""
            ev[c] = raw

        stmt_key, orig_idx = row_mapping[i] if i < len(row_mapping) else ("", 0)

        # --- Separator row ---
        if stmt_key == "__separator__":
            rows.append({
                "seg": 0,
                "ev": [ev.get(c, "") for c in cols],
                "pv": [ev.get(c, "") for c in cols],
                "ok": True,
                "dc": [],
                "stmtKey": "__separator__",
                "pos": None,
                "rowType": "separator",
            })
            continue

        # --- Opening Balance row ---
        if orig_idx == -1:
            rows.append({
                "seg": 0,
                "ev": [ev.get(c, "") for c in cols],
                "pv": [ev.get(c, "") for c in cols],
                "ok": True,
                "dc": [],
                "stmtKey": stmt_key,
                "pos": None,
                "rowType": "opening_balance",
            })
            continue

        # --- Transaction row ---
        pv = dict(ev)
        vf = verify_files.get(stmt_key)
        if vf and vf.get("expected_dfs"):
            pdf_df = vf["expected_dfs"][0]
            if orig_idx < len(pdf_df):
                for c in cols:
                    if c == "Statement":
                        continue
                    if c in pdf_df.columns:
                        raw_p = str(pdf_df.iloc[orig_idx][c])
                        pv[c] = "" if raw_p == "nan" else raw_p

        # Diff detection (skip Statement column)
        diff_cols = []
        for ci, c in enumerate(cols):
            if c == "Statement":
                continue
            if _normalize(pv.get(c, "")) != _normalize(ev.get(c, "")):
                diff_cols.append(ci)

        # PDF position
        pos = None
        if vf:
            positions = vf.get("row_positions", [])
            if orig_idx < len(positions):
                rp = positions[orig_idx]
                pos = {
                    "pg": rp.page_index,
                    "yt": round(rp.y_top, 1),
                    "yb": round(rp.y_bottom, 1),
                    "ph": round(rp.page_height, 1),
                }

        tx_total += 1
        rows.append({
            "seg": 0,
            "ev": [ev.get(c, "") for c in cols],
            "pv": [pv.get(c, "") for c in cols],
            "ok": len(diff_cols) == 0,
            "dc": diff_cols,
            "stmtKey": stmt_key,
            "pos": pos,
            "rowType": "transaction",
        })

    tx_matched = sum(1 for r in rows if r.get("rowType") == "transaction" and r["ok"])
    tx_mismatched = sum(1 for r in rows if r.get("rowType") == "transaction" and not r["ok"])

    return {
        "isMerged": True,
        "segments": segments,
        "rows": rows,
        "total": len(rows),
        "txTotal": tx_total,
        "matched": tx_matched,
        "mismatched": tx_mismatched,
    }


def _build_merged_verification_html(payload_json: str, pdfs_map_json: str) -> str:
    """Build HTML for the merged multi-PDF verification component."""
    tpl = (Path(__file__).parent / "templates" / "verification_merged.html").read_text(encoding="utf-8")
    return tpl.replace('"__PAYLOAD__"', payload_json).replace('"__PDFS_MAP__"', pdfs_map_json)


# --- 数据复核页面 / Verification Page ---
if st.session_state.get("show_verification"):
    # 全屏模式：隐藏侧边栏
    st.markdown(
        """<style>
        [data-testid="stSidebar"] { display: none; }
        [data-testid="stSidebarCollapsedControl"] { display: none; }
        .block-container { max-width: 100% !important; padding-left: 1rem; padding-right: 1rem; }
        </style>""",
        unsafe_allow_html=True,
    )

    _vdata = st.session_state.get("verification_data")
    if not _vdata:
        st.warning(
            "暂无可复核数据，请先上传 PDF 并完成批量转换。\n\n"
            "No data to verify. Please upload PDF and run batch conversion first."
        )
        if st.button("返回 / Back", key="verify_back_empty"):
            st.session_state.show_verification = False
            st.rerun()
        st.stop()

    # === Merged mode verification ===
    _merged_data = _vdata.get("merged")
    if _merged_data and _merged_data.get("is_merged"):
        _m_hdr = st.columns([5, 2, 1.5])
        with _m_hdr[0]:
            st.subheader("🔍 合并数据复核 / Merged Data Verification")
        with _m_hdr[1]:
            st.download_button(
                "⬇ 下载合并 Excel / Download",
                _merged_data["merged_excel_bytes"],
                file_name=f"{_merged_data['merged_filename']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="verify_download_merged",
            )
        with _m_hdr[2]:
            if st.button("↩ 返回 / Back", key="verify_back_merged", use_container_width=True):
                st.session_state.show_verification = False
                st.rerun()

        # Build merged payload
        _m_payload = _build_merged_verification_payload(
            _merged_data["merged_excel_bytes"],
            _merged_data["row_mapping"],
            _vdata["files"],
        )
        _m_payload_json = json.dumps(_m_payload, ensure_ascii=False)

        # Build PDFs map (stem → base64)
        _m_pdfs_map = {}
        for _stem, _pdf_bytes in _merged_data["pdf_map"].items():
            _m_pdfs_map[_stem] = base64.b64encode(_pdf_bytes).decode()
        _m_pdfs_json = json.dumps(_m_pdfs_map, ensure_ascii=False)

        # Render merged verification component
        _m_html = _build_merged_verification_html(_m_payload_json, _m_pdfs_json)
        stc.html(_m_html, height=720, scrolling=False)

        st.divider()
        _m_tx_total = _m_payload.get("txTotal", 0)
        _m_mismatched = _m_payload["mismatched"]
        if _m_mismatched == 0:
            st.success(f"✅ 全部 {_m_tx_total} 条交易核对一致 / All {_m_tx_total} transactions matched")
        else:
            st.warning(f"⚠️ {_m_tx_total} 条交易中发现 {_m_mismatched} 处不一致 / Found {_m_mismatched} mismatch(es) in {_m_tx_total} transactions")

        st.stop()

    # --- Header row: title | file selector | download | back ---
    _v_file_names = list(_vdata["files"].keys())
    _hdr = st.columns([3.5, 3, 1.5, 1.5])
    with _hdr[0]:
        st.subheader("🔍 数据复核 / Data Verification")
    with _hdr[1]:
        _v_sel_file = st.selectbox(
            "文件 / File", _v_file_names, key="verify_file_sel",
        )
    with _hdr[2]:
        st.write("")  # vertical spacer to align with selectbox
        _v_dl_info = _vdata["files"][_v_sel_file]
        st.download_button(
            "⬇ 下载 / Download",
            _v_dl_info["excel_bytes"],
            file_name=f"{_v_sel_file}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="verify_download_excel",
        )
    with _hdr[3]:
        st.write("")  # vertical spacer
        if st.button("↩ 返回 / Back", key="verify_back", use_container_width=True):
            st.session_state.show_verification = False
            st.rerun()

    _v_info = _vdata["files"][_v_sel_file]
    _v_result = _v_info["result"]
    _v_pdf_bytes = _v_info["pdf_bytes"]
    _v_expected_dfs = _v_info["expected_dfs"]
    _v_excel_bytes = _v_info["excel_bytes"]
    pdf_b64 = base64.b64encode(_v_pdf_bytes).decode()

    # 读取 Excel 全部 sheet
    _v_excel_dfs: dict[str, pd.DataFrame] = {}
    xls = pd.ExcelFile(io.BytesIO(_v_excel_bytes))
    for sn in xls.sheet_names:
        _v_excel_dfs[sn] = pd.read_excel(xls, sheet_name=sn)

    # --- 构建统一行列表（全部 sheet 合并） ---
    # 每行: {sheet, row, pdf_vals:{col:str}, excel_vals:{col:str}, columns:[], diff_cols:set}
    _unified_rows: list[dict] = []

    # 差异 lookup: (sheet, row, col)
    _diff_set: set[tuple[str, int, str]] = set()
    for d in _v_result.diffs:
        _diff_set.add((d.sheet, d.row, d.column))

    # Summary PDF 原始值 (从 StatementInfo 解析)
    _summary_pdf_vals_map: dict[str, str] = _v_info.get("summary_pdf_vals", {})

    if _v_info.get("is_statement"):
        # 银行账单: Summary + Transactions
        if "Summary" in _v_excel_dfs:
            sdf = _v_excel_dfs["Summary"]
            for r in range(len(sdf)):
                cols = list(sdf.columns)
                # Excel 值
                ev = {c: str(sdf.iloc[r][c]) if str(sdf.iloc[r][c]) != "nan" else "" for c in cols}
                # PDF 值: 从 Excel 的 Item 列读取 item 名称，查找 PDF 解析值
                pv = dict(ev)  # 默认与 Excel 相同
                if "Item" in cols and "Value" in cols:
                    excel_item = ev.get("Item", "")
                    pdf_val = _summary_pdf_vals_map.get(excel_item, "")
                    pv["Value"] = pdf_val
                # 比较差异
                diff_cols: set[str] = set()
                for c in cols:
                    if _normalize(pv.get(c, "")) != _normalize(ev.get(c, "")):
                        diff_cols.add(c)
                _unified_rows.append({
                    "sheet": "Summary", "row": r, "columns": cols,
                    "pdf_vals": pv, "excel_vals": ev,
                    "diff_cols": diff_cols,
                })
        if "Transactions" in _v_excel_dfs:
            edf = _v_excel_dfs["Transactions"]
            pdf_df = _v_expected_dfs[0] if _v_expected_dfs else edf
            cols = list(edf.columns)
            max_r = max(len(pdf_df), len(edf))
            for r in range(max_r):
                pv = {}
                ev = {}
                for c in cols:
                    raw_p = str(pdf_df.iloc[r][c]) if r < len(pdf_df) and c in pdf_df.columns else ""
                    pv[c] = "" if raw_p == "nan" else raw_p
                    raw_e = str(edf.iloc[r][c]) if r < len(edf) else ""
                    ev[c] = "" if raw_e == "nan" else raw_e
                _unified_rows.append({
                    "sheet": "Transactions", "row": r, "columns": cols,
                    "pdf_vals": pv, "excel_vals": ev,
                    "diff_cols": {c for c in cols if ("Transactions", r, c) in _diff_set},
                })
    else:
        # 标准表格模式
        for i, exp_df in enumerate(_v_expected_dfs):
            sn = f"Table_{i + 1}"
            edf = _v_excel_dfs.get(sn, pd.DataFrame())
            cols = list(edf.columns) if not edf.empty else list(exp_df.columns)
            max_r = max(len(exp_df), len(edf))
            for r in range(max_r):
                pv = {}
                ev = {}
                for c in cols:
                    raw_p = str(exp_df.iloc[r][c]) if r < len(exp_df) and c in exp_df.columns else ""
                    pv[c] = "" if raw_p == "nan" else raw_p
                    raw_e = str(edf.iloc[r][c]) if r < len(edf) else ""
                    ev[c] = "" if raw_e == "nan" else raw_e
                _unified_rows.append({
                    "sheet": sn, "row": r, "columns": cols,
                    "pdf_vals": pv, "excel_vals": ev,
                    "diff_cols": {c for c in cols if (sn, r, c) in _diff_set},
                })

    # --- 获取行坐标和表头 ---
    _row_positions: list[RowPosition] = _v_info.get("row_positions", [])
    _row_headers: list[str] = _v_info.get("row_headers", [])
    _is_statement = _v_info.get("is_statement", False)

    # --- 获取 Summary PDF 坐标 ---
    _summary_positions: list | None = None
    if _is_statement and "Summary" in _v_excel_dfs:
        _sdf = _v_excel_dfs["Summary"]
        _sum_items = []
        for _r in range(len(_sdf)):
            _item = str(_sdf.iloc[_r].get("Item", "")) if "Item" in _sdf.columns else ""
            _val = str(_sdf.iloc[_r].get("Value", "")) if "Value" in _sdf.columns else ""
            if _item == "nan":
                _item = ""
            if _val == "nan":
                _val = ""
            _sum_items.append({"item": _item, "value": _val})
        try:
            _summary_positions = BankStatementParser().get_summary_positions(_v_pdf_bytes, _sum_items)
        except Exception:
            _summary_positions = None

    # Extract layers if TripleVerificationResult
    _v_layers = getattr(_v_result, "layers", None)

    # Build payload for dual-panel verification component
    _payload = _build_verification_payload(
        _unified_rows, _row_positions, _row_headers, _is_statement,
        summary_positions=_summary_positions,
        layers=_v_layers,
    )
    _payload_json = json.dumps(_payload, ensure_ascii=False)

    # --- 双面板核对组件 / Dual-panel verification component ---
    _html = _build_verification_html(_payload_json, pdf_b64)
    stc.html(_html, height=720, scrolling=False)

    st.divider()

    # 验证结果汇总 — 三层状态
    if _v_layers:
        _layer_cols = st.columns(len(_v_layers) + 1)
        for _li, _layer in enumerate(_v_layers):
            with _layer_cols[_li]:
                if _layer.skipped:
                    st.info(f"⏭ {_layer.label}: 跳过 / Skipped")
                elif _layer.passed:
                    st.success(f"✅ {_layer.label}")
                else:
                    st.error(f"❌ {_layer.label} ({_layer.mismatched_cells})")
        with _layer_cols[-1]:
            if _v_result.matched:
                st.success(f"✅ {_v_result.message}")
            else:
                st.warning(f"⚠️ {_v_result.message}")

        # Arithmetic issues detail
        _arith_layer = next((l for l in _v_layers if l.name == "arithmetic"), None)
        if _arith_layer and _arith_layer.arithmetic_issues:
            with st.expander(f"🔢 算术校验详情 / Arithmetic Details ({len(_arith_layer.arithmetic_issues)})"):
                for _ai in _arith_layer.arithmetic_issues:
                    st.warning(f"• {_ai.message}")
    else:
        if _v_result.matched:
            st.success(f"✅ {_v_result.message}")
        else:
            st.warning(f"⚠️ {_v_result.message}")

    # 差异列表
    if _v_result.diffs:
        st.markdown("### 📋 差异列表 / Difference List")
        diff_table = pd.DataFrame([
            {
                "Sheet": d.sheet,
                "行 / Row": d.row + 1,
                "列 / Column": d.column,
                "PDF 值 / Expected": d.expected or "(空/empty)",
                "Excel 值 / Actual": d.actual or "(空/empty)",
            }
            for d in _v_result.diffs
        ])
        st.dataframe(diff_table, use_container_width=True)

    # 可编辑版本
    with st.expander("✏️ 手工修正 / Manual Edit"):
        _edit_sheet = st.selectbox("Sheet", list(_v_excel_dfs.keys()), key="verify_edit_sheet")
        _edit_df = _v_excel_dfs[_edit_sheet]
        edited_df = st.data_editor(
            _edit_df, use_container_width=True, num_rows="dynamic",
            key=f"verify_editor_{_v_sel_file}_{_edit_sheet}",
        )
        if st.button("💾 重新导出 / Re-export Excel", key="verify_reexport"):
            out_buf = io.BytesIO()
            with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
                for sn in _v_excel_dfs:
                    df_w = edited_df if sn == _edit_sheet else _v_excel_dfs[sn]
                    df_w.to_excel(writer, sheet_name=sn, index=False)
            st.download_button(
                "下载修正后的 Excel / Download Corrected Excel",
                out_buf.getvalue(),
                file_name=f"{_v_sel_file}_corrected.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="verify_download_corrected",
            )

    st.stop()

# --- 文件上传 / File Upload ---
uploaded_files = st.file_uploader(
    "上传 PDF 文件 / Upload PDF Files",
    type=["pdf"],
    accept_multiple_files=True,
)

if not uploaded_files:
    # No files → clear stale verification cache
    st.session_state.pop("verification_data", None)
    st.session_state.pop("_uploaded_file_names", None)
    st.info("请上传一个或多个 PDF 文件开始转换。/ Please upload one or more PDF files to start.")
    st.stop()

# Track uploaded file names; clear verification cache when files change
_cur_file_names = frozenset(f.name for f in uploaded_files)
_prev_file_names = st.session_state.get("_uploaded_file_names", frozenset())
if _cur_file_names != _prev_file_names:
    st.session_state.pop("verification_data", None)
    st.session_state["_uploaded_file_names"] = _cur_file_names

st.divider()

extractor = PDFExtractor()
converter = ExcelConverter()
statement_parser = BankStatementParser()


def _save_to_tmp(uf) -> str:
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(uf.getvalue())
        return tmp.name


def _try_extract(tmp_path: str, mode: str):
    """尝试提取 PDF 内容。返回 (tables, statement_info)。"""
    if mode == "银行账单":
        info = statement_parser.parse(tmp_path)
        return [], info

    result = extractor.extract(tmp_path)
    if result.tables or mode == "标准表格":
        return result.tables, None

    # 自动检测：标准提取无结果，回退到银行账单解析
    info = statement_parser.parse(tmp_path)
    if not info.transactions.empty:
        return [], info
    return [], None


def _statement_to_excel(info, output_path: Path) -> Path:
    """将银行账单信息写入 Excel。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_data = {
            "Item": ["Bank Name", "Account Number", "Period From", "Period To",
                     "Opening Balance", "Closing Balance",
                     "Total Deposits", "Total Withdrawals"],
            "Value": [info.bank_name, info.account_number, info.period_from, info.period_to,
                      info.opening_balance, info.closing_balance,
                      info.total_deposits, info.total_withdrawals],
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        info.transactions.to_excel(writer, sheet_name="Transactions", index=False)

        for sheet_name in ["Summary", "Transactions"]:
            ws = writer.sheets[sheet_name]
            df = summary_df if sheet_name == "Summary" else info.transactions
            for col_idx, col_name in enumerate(df.columns, 1):
                max_len = len(str(col_name))
                for val in df.iloc[:, col_idx - 1]:
                    max_len = max(max_len, len(str(val)) if val is not None else 0)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 55)
    return output_path


def _extract_year_from_period(period_str: str) -> int | None:
    """Extract 4-digit year from period string like 'November 1, 2024'."""
    if not period_str:
        return None
    m = re.search(r"(20\d{2})", period_str)
    return int(m.group(1)) if m else None


def _extract_month_from_period(period_str: str) -> int | None:
    """Extract month number from period string like 'January 1, 2025'."""
    if not period_str:
        return None
    dt = pd.to_datetime(period_str, errors="coerce")
    return dt.month if pd.notna(dt) else None


def _parse_tx_month(date_str: str) -> int | None:
    """Parse month from transaction date like 'Jan 03' or 'Dec 28'."""
    if not date_str or not str(date_str).strip():
        return None
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    m = re.match(r"([A-Za-z]{3})", str(date_str).strip())
    if m:
        return month_map.get(m.group(1).lower())
    return None


def _filter_january_transactions(
    txn_df: pd.DataFrame,
    period_year: int,
    majority_year: int,
) -> pd.DataFrame:
    """Filter transactions in a January PDF based on year rules.

    - January PDF of majority year: remove Dec transactions (prior year)
    - January PDF of majority year + 1: remove Jan transactions, keep Dec (majority year)
    """
    if txn_df.empty:
        return txn_df

    tx_months = txn_df["Date"].apply(_parse_tx_month)

    if period_year == majority_year:
        # Jan of majority year: remove Dec rows (they belong to prior year)
        keep = tx_months != 12
    elif period_year == majority_year + 1:
        # Jan of next year: remove Jan rows, keep Dec rows (majority year)
        keep = tx_months == 12
    else:
        # Other years: remove all (shouldn't normally happen)
        return txn_df.iloc[0:0]

    return txn_df[keep]  # preserve original index for row_mapping


def _build_merged_transactions(
    all_txns: list[tuple[str, pd.DataFrame, object]],
    output_dir: Path,
) -> tuple[Path, list[tuple[str, int]]] | None:
    """Merge statement transactions into block-based Excel with year filtering.

    Output structure: blocks sorted by period_from (oldest first), each block
    starts with an Opening Balance row, preserves original PDF row order,
    separated by 2 empty rows between blocks.

    Args:
        all_txns: [(pdf_stem, transactions_df, stmt_info), ...]
        output_dir: output directory
    Returns:
        (path_to_excel, row_mapping) or None.
        row_mapping: [(stem, orig_idx)] per output row.
          orig_idx = -1 for Opening Balance rows,
          stem = "__separator__" for empty separator rows.
    """
    if not all_txns:
        return None

    # --- Collect metadata ---
    bank_name = ""
    account = ""
    entries: list[dict] = []

    for pdf_stem, txn_df, info in all_txns:
        period_from_dt = pd.to_datetime(info.period_from, errors="coerce") if info.period_from else pd.NaT
        period_to_dt = pd.to_datetime(info.period_to, errors="coerce") if info.period_to else pd.NaT
        period_year = _extract_year_from_period(info.period_from) or _extract_year_from_period(info.period_to)
        period_month = _extract_month_from_period(info.period_from)

        entries.append({
            "stem": pdf_stem,
            "txn_df": txn_df,
            "info": info,
            "period_from_dt": period_from_dt,
            "period_year": period_year or pd.Timestamp.now().year,
            "period_month": period_month,
        })

        if info.bank_name and not bank_name:
            bank_name = info.bank_name
        if info.account_number and not account:
            account = info.account_number

    # --- Determine majority year ---
    year_counts: dict[int, int] = {}
    for e in entries:
        y = e["period_year"]
        year_counts[y] = year_counts.get(y, 0) + 1
    majority_year = max(year_counts, key=year_counts.get)

    # --- Sort entries by period_from (oldest first) ---
    entries.sort(key=lambda e: e["period_from_dt"] if pd.notna(e["period_from_dt"]) else pd.Timestamp.max)

    # --- Build blocks ---
    cols = ["Statement", "Date", "Description", "Withdrawals", "Deposits", "Balance"]
    all_rows: list[dict] = []
    row_mapping: list[tuple[str, int]] = []
    blocks_added = 0

    for entry in entries:
        stem = entry["stem"]
        info = entry["info"]
        txn_df = entry["txn_df"].copy()
        period_month = entry["period_month"]
        period_year = entry["period_year"]

        # --- Year filtering (January PDFs only) ---
        if period_month == 1:
            txn_df = _filter_january_transactions(txn_df, period_year, majority_year)

        # Skip empty blocks after filtering
        if txn_df.empty:
            continue

        # --- Separator between blocks ---
        if blocks_added > 0:
            for _ in range(2):
                all_rows.append({c: "" for c in cols})
                row_mapping.append(("__separator__", -1))

        # --- Opening Balance row ---
        ob_row = {c: "" for c in cols}
        ob_row["Statement"] = stem
        ob_row["Date"] = info.period_from or ""
        ob_row["Description"] = "Opening Balance"
        ob_row["Balance"] = info.opening_balance or ""
        all_rows.append(ob_row)
        row_mapping.append((stem, -1))

        # --- Transaction rows (preserve original order, skip duplicate Opening Balance) ---
        for orig_idx in range(len(txn_df)):
            desc = str(txn_df.iloc[orig_idx].get("Description", "")) if "Description" in txn_df.columns else ""
            if re.search(r"Opening\s*Balance", desc, re.I):
                continue
            row_data = {c: "" for c in cols}
            row_data["Statement"] = stem
            for c in cols:
                if c == "Statement":
                    continue
                if c in txn_df.columns:
                    val = str(txn_df.iloc[orig_idx][c])
                    row_data[c] = "" if val == "nan" else val
            all_rows.append(row_data)
            # txn_df preserves original index after filtering, so
            # txn_df.index[i] gives the original row number for position mapping
            row_mapping.append((stem, int(txn_df.index[orig_idx])))

        blocks_added += 1

    if not all_rows:
        return None

    merged = pd.DataFrame(all_rows, columns=cols)

    # --- Filename ---
    bank_short = bank_name.replace(" ", "") if bank_name else "Bank"
    last4 = account[-4:] if len(account) >= 4 else (account or "0000")
    period_from_dates = [e["period_from_dt"] for e in entries if pd.notna(e["period_from_dt"])]
    period_to_dates = [
        pd.to_datetime(e["info"].period_to, errors="coerce")
        for e in entries
        if pd.notna(pd.to_datetime(e["info"].period_to, errors="coerce"))
    ]
    d_start = min(period_from_dates).strftime("%Y%m%d") if period_from_dates else "start"
    d_end = max(period_to_dates).strftime("%Y%m%d") if period_to_dates else "end"
    filename = f"{bank_short}_{last4}_{d_start}-{d_end}.xlsx"

    out_path = output_dir / filename
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Transactions", index=False)
        ws = writer.sheets["Transactions"]
        for col_idx, col_name in enumerate(merged.columns, 1):
            max_len = len(str(col_name))
            for val in merged.iloc[:, col_idx - 1]:
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 55)

    return out_path, row_mapping


# --- 预览 & 转换 / Preview & Convert ---
tabs = st.tabs(["📋 预览 / Preview", "⚡ 批量转换 / Batch Convert"])

with tabs[0]:
    st.subheader("表格预览 / Table Preview")
    for uf in uploaded_files:
        with st.expander(f"📄 {uf.name}", expanded=len(uploaded_files) == 1):
            tmp_path = _save_to_tmp(uf)
            try:
                tables, stmt_info = _try_extract(tmp_path, parse_mode)

                if tables:
                    st.success(
                        f"检测到 {len(tables)} 个表格 / {len(tables)} table(s) detected"
                    )
                    for i, df in enumerate(tables):
                        st.markdown(f"**表格 / Table {i + 1}**")
                        st.dataframe(df, use_container_width=True)
                elif stmt_info and not stmt_info.transactions.empty:
                    st.success(
                        f"银行账单模式 — 检测到 {len(stmt_info.transactions)} 笔交易 / "
                        f"Bank statement — {len(stmt_info.transactions)} transaction(s)"
                    )
                    st.markdown("**账户摘要 / Account Summary**")
                    if stmt_info.bank_name:
                        st.metric("银行 / Bank", stmt_info.bank_name)
                    col1, col2 = st.columns(2)
                    col1.metric("账户 / Account", stmt_info.account_number)
                    col2.metric(
                        "期间 / Period",
                        f"{stmt_info.period_from} ~ {stmt_info.period_to}",
                    )
                    col3, col4 = st.columns(2)
                    col3.metric("总存入 / Deposits", f"${stmt_info.total_deposits}")
                    col4.metric("总支出 / Withdrawals", f"${stmt_info.total_withdrawals}")
                    st.markdown("**交易明细 / Transactions**")
                    st.dataframe(stmt_info.transactions, use_container_width=True)
                else:
                    st.warning("未检测到表格或交易数据。/ No tables or transactions detected.")
            except Exception as e:
                st.error(f"解析失败 / Parse failed: {e}")

with tabs[1]:
    st.subheader("批量转换 / Batch Convert")
    st.write(f"已上传 **{len(uploaded_files)}** 个文件 / file(s)")

    if st.button("开始转换 / Start", type="primary", use_container_width=True):
        # 50MB size limit for merge mode
        if not sheet_per_table:
            _total_pdf_size = sum(len(uf.getvalue()) for uf in uploaded_files)
            _MAX_MERGE_SIZE = 50 * 1024 * 1024
            if _total_pdf_size > _MAX_MERGE_SIZE:
                st.error(
                    f"⚠️ 合并模式下 PDF 总大小不可超过 50MB（当前 {_total_pdf_size / 1024 / 1024:.1f}MB）。\n\n"
                    f"请使用 [PDF24](https://www.pdf24.org) 等工具压缩 PDF 后重新上传。\n\n"
                    f"Total PDF size exceeds 50MB limit for merge mode "
                    f"({_total_pdf_size / 1024 / 1024:.1f}MB). "
                    f"Please compress your PDFs using PDF24 or similar tools before uploading."
                )
                st.stop()

        progress_bar = st.progress(0, text="准备中... / Preparing...")
        _verifier = DataVerifier()
        _triple_verifier = TripleVerifier()
        _text_extractor = TextBasedExtractor()

        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "input"
            output_dir = Path(tmpdir) / "output"
            input_dir.mkdir()
            output_dir.mkdir()

            pdf_paths = []
            uf_map: dict[str, bytes] = {}
            for uf in uploaded_files:
                p = input_dir / uf.name
                pdf_bytes = uf.getvalue()
                p.write_bytes(pdf_bytes)
                pdf_paths.append(p)
                uf_map[uf.name] = pdf_bytes

            total = len(pdf_paths)
            succeeded = 0
            failed = 0
            errors: list[str] = []
            verify_files: dict = {}
            merge_txn_data: list = []

            for i, pdf_path in enumerate(pdf_paths):
                try:
                    tables, stmt_info = _try_extract(str(pdf_path), parse_mode)
                    out_name = pdf_path.stem + ".xlsx"
                    out_path = output_dir / out_name

                    if stmt_info and not stmt_info.transactions.empty:
                        _statement_to_excel(stmt_info, out_path)
                        # 获取行坐标
                        row_headers, row_positions = statement_parser.get_row_positions(pdf_path)
                        # Summary fields for arithmetic checks
                        summary_dict = {
                            "opening_balance": stmt_info.opening_balance,
                            "closing_balance": stmt_info.closing_balance,
                            "total_deposits": stmt_info.total_deposits,
                            "total_withdrawals": stmt_info.total_withdrawals,
                        }
                        # Source B: text-based extraction (with fallback)
                        source_b_df = pd.DataFrame()
                        try:
                            text_result = _text_extractor.extract(pdf_path)
                            source_b_df = text_result.transactions
                        except Exception:
                            pass  # Layer 1 will be marked as skipped
                        # Triple verification
                        v_result = _triple_verifier.verify_statement(
                            source_a=stmt_info.transactions,
                            source_b=source_b_df,
                            excel_path=out_path,
                            summary=summary_dict,
                        )
                        # Summary 字段: PDF 解析的原始值
                        summary_pdf_vals = {
                            "Bank Name": stmt_info.bank_name,
                            "Account Number": stmt_info.account_number,
                            "Period From": stmt_info.period_from,
                            "Period To": stmt_info.period_to,
                            "Opening Balance": stmt_info.opening_balance,
                            "Closing Balance": stmt_info.closing_balance,
                            "Total Deposits": stmt_info.total_deposits,
                            "Total Withdrawals": stmt_info.total_withdrawals,
                        }
                        verify_files[pdf_path.stem] = {
                            "pdf_bytes": uf_map[pdf_path.name],
                            "expected_dfs": [stmt_info.transactions],
                            "excel_bytes": out_path.read_bytes(),
                            "result": v_result,
                            "is_statement": True,
                            "row_positions": row_positions,
                            "row_headers": row_headers,
                            "summary_pdf_vals": summary_pdf_vals,
                            "source_b_df": source_b_df,
                        }
                        if not sheet_per_table:
                            merge_txn_data.append((pdf_path.stem, stmt_info.transactions, stmt_info))
                    elif tables:
                        from core.extractor import ExtractionResult
                        result = ExtractionResult(tables=tables, page_count=0, source=str(pdf_path))
                        converter.convert(result, out_path, sheet_per_table)
                        # 自动验证
                        v_result = _verifier.verify(tables, out_path)
                        verify_files[pdf_path.stem] = {
                            "pdf_bytes": uf_map[pdf_path.name],
                            "expected_dfs": tables,
                            "excel_bytes": out_path.read_bytes(),
                            "result": v_result,
                            "is_statement": False,
                        }
                    else:
                        errors.append(f"{pdf_path.name}: 未检测到数据 / No data detected")
                        failed += 1
                        continue
                    succeeded += 1
                except Exception as e:
                    errors.append(f"{pdf_path.name}: {e}")
                    failed += 1

                progress_bar.progress((i + 1) / total, text=f"处理中... / Processing {i + 1}/{total}")

            # Merge mode: combine all statement transactions
            merged_excel_path: Path | None = None
            merged_row_mapping: list[tuple[str, int]] = []
            if not sheet_per_table and merge_txn_data:
                result = _build_merged_transactions(merge_txn_data, output_dir)
                if result:
                    merged_excel_path, merged_row_mapping = result

            progress_bar.progress(1.0, text="完成！/ Done!")

            # 存储验证数据到 session_state
            if verify_files:
                vdata: dict = {"files": verify_files}
                # Store merged verification data
                if merged_excel_path and merged_row_mapping:
                    vdata["merged"] = {
                        "is_merged": True,
                        "merged_excel_bytes": merged_excel_path.read_bytes(),
                        "merged_filename": merged_excel_path.stem,
                        "row_mapping": merged_row_mapping,
                        "pdf_map": {stem: vf["pdf_bytes"] for stem, vf in verify_files.items()},
                        "positions_map": {stem: vf.get("row_positions", []) for stem, vf in verify_files.items()},
                        "headers_map": {stem: vf.get("row_headers", []) for stem, vf in verify_files.items()},
                    }
                st.session_state.verification_data = vdata

            col1, col2, col3 = st.columns(3)
            col1.metric("总计 / Total", total)
            col2.metric("成功 / Success", succeeded)
            col3.metric("失败 / Failed", failed)

            for err in errors:
                st.error(f"❌ {err}")

            # 自动验证结果汇总
            if verify_files:
                st.divider()
                all_matched = all(v["result"].matched for v in verify_files.values())
                total_diffs = sum(v["result"].mismatched_cells for v in verify_files.values())
                if all_matched:
                    st.success(
                        "✅ 三源交叉验证全部通过 / "
                        "All three verification layers passed"
                    )
                else:
                    st.warning(
                        f"⚠️ 发现 {total_diffs} 处不一致，请点击侧边栏「数据复核」查看详情 / "
                        f"Found {total_diffs} mismatch(es), click 'Data Verification' in sidebar for details"
                    )

            if merged_excel_path:
                with open(merged_excel_path, "rb") as f:
                    st.download_button(
                        "下载合并 Excel / Download Merged",
                        f.read(),
                        file_name=merged_excel_path.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            else:
                excel_files = list(output_dir.glob("*.xlsx"))
                if excel_files:
                    if len(excel_files) == 1:
                        with open(excel_files[0], "rb") as f:
                            st.download_button(
                                "下载 Excel / Download",
                                f.read(),
                                file_name=excel_files[0].name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                            )
                    else:
                        buf = io.BytesIO()
                        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                            for ef in excel_files:
                                zf.write(ef, ef.name)
                        st.download_button(
                            f"下载全部 / Download All ({len(excel_files)} files, ZIP)",
                            buf.getvalue(),
                            file_name="pdf2excel_output.zip",
                            mime="application/zip",
                            use_container_width=True,
                        )
