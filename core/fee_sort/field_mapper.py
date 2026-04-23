"""Map merged Excel rows to StandardRow objects for classification."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

import pandas as pd


@dataclass
class StandardRow:
    """Normalized transaction row ready for rule-engine classification.

    For bank statements: withdrawals/deposits are split (always non-negative floats).
    For credit cards: the signed Amount is stored in `withdrawals` (withdrawals_float
    can be negative for refunds), and deposits is unused. `schema` selects the
    appropriate downstream handling.
    """

    idx: int
    row_type: str  # "separator" | "opening_balance" | "transaction"
    statement: str
    date: str
    description: str
    withdrawals: str
    deposits: str
    balance: str
    withdrawals_float: float = 0.0
    deposits_float: float = 0.0
    schema: str = "bank"  # "bank" | "credit_card"


def _to_float(val: str) -> float:
    """Parse a currency string to float, stripping commas / whitespace."""
    if not val:
        return 0.0
    cleaned = re.sub(r"[^\d.\-]", "", str(val))
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _classify_row_type(
    stem: str,
    orig_idx: int,
    description: str,
) -> str:
    """Determine row type from row_mapping + description."""
    if stem == "__separator__":
        return "separator"
    if orig_idx == -1:
        # Opening Balance (bank) and Previous Balance (credit card) both
        # map to the "opening_balance" type — P0 excludes them identically.
        return "opening_balance"
    return "transaction"


class FieldMapper:
    """Convert merged Excel bytes + row_mapping → list[StandardRow]."""

    @staticmethod
    def map(
        merged_excel_bytes: bytes,
        row_mapping: list[tuple[str, int]],
    ) -> list[StandardRow]:
        """Read merged Excel and return a StandardRow per data row.

        Args:
            merged_excel_bytes: raw .xlsx bytes from merged conversion.
            row_mapping: [(pdf_stem, orig_row_idx)] parallel to Excel data rows.
                stem="__separator__" for blank separator rows,
                orig_idx=-1 for Opening Balance rows.
        Returns:
            List of StandardRow, one per data row.
        """
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(merged_excel_bytes), read_only=True, data_only=True)
        ws = wb.active
        # Read header row to build column index
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map: dict[str, int] = {}
        for ci, name in enumerate(header_row):
            if name:
                col_map[str(name)] = ci

        # Read all data rows (row 2 onward), preserving blanks
        data_rows: list[tuple] = []
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            data_rows.append(row_vals)
        wb.close()

        # openpyxl truncates trailing all-empty rows. If row_mapping expects
        # more rows (typically separator rows at the tail), pad with blanks.
        empty_row = tuple(None for _ in range(len(col_map) or len(header_row)))
        while len(data_rows) < len(row_mapping):
            data_rows.append(empty_row)

        def _cell(row_vals: tuple, col_name: str) -> str:
            idx = col_map.get(col_name)
            if idx is None or idx >= len(row_vals):
                return ""
            v = row_vals[idx]
            return str(v) if v is not None else ""

        # Detect schema: credit-card merged Excel has "Amount" and lacks the
        # bank schema columns. Map Amount → withdrawals (positive) or
        # deposits (negative, abs value) so rule engine can reuse its logic.
        is_credit_card = (
            "Amount" in col_map
            and "Withdrawals" not in col_map
            and "Deposits" not in col_map
        )

        rows: list[StandardRow] = []
        for i, (stem, orig_idx) in enumerate(row_mapping):
            if i >= len(data_rows):
                break

            rv = data_rows[i]
            statement = _cell(rv, "Statement")

            if is_credit_card:
                date = _cell(rv, "Transaction Date")
                description = _cell(rv, "Activity Description")
                amount_str = _cell(rv, "Amount")
                # Preserve the sign: positive = charge, negative = refund/payment.
                # The downstream rule engine (CreditCardRuleEngine) distinguishes
                # excluded rows (PAYMENT / CASH BACK REWARD) from refunds based
                # on description pattern + sign.
                amount_clean = amount_str.replace("$", "").strip()
                withdrawals = amount_clean
                deposits = ""
                withdrawals_float = _to_float(amount_str)  # signed
                deposits_float = 0.0
                balance = ""
                schema = "credit_card"
            else:
                date = _cell(rv, "Date")
                description = _cell(rv, "Description")
                withdrawals = _cell(rv, "Withdrawals")
                deposits = _cell(rv, "Deposits")
                balance = _cell(rv, "Balance")
                withdrawals_float = _to_float(withdrawals)
                deposits_float = _to_float(deposits)
                schema = "bank"

            row_type = _classify_row_type(stem, orig_idx, description)

            rows.append(
                StandardRow(
                    idx=i,
                    row_type=row_type,
                    statement=statement,
                    date=date,
                    description=description,
                    withdrawals=withdrawals,
                    deposits=deposits,
                    balance=balance,
                    withdrawals_float=withdrawals_float,
                    deposits_float=deposits_float,
                    schema=schema,
                )
            )

        return rows
