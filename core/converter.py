from __future__ import annotations

import re
from pathlib import Path

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core.extractor import ExtractionResult


def right_align_numbers(ws) -> None:
    """Right-align all numeric cells in a worksheet (skip header row)."""
    _NUM_RE = re.compile(r"^-?[\d,]+\.?\d*$")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val and _NUM_RE.match(val):
                    cell.alignment = Alignment(horizontal="right")


class ExcelConverter:
    """将 ExtractionResult 转为 Excel 文件。"""

    def convert(
        self,
        result: ExtractionResult,
        output: str | Path,
        sheet_per_table: bool = True,
    ) -> Path:
        """将提取结果写入 Excel。

        Args:
            result: 提取结果
            output: 输出文件路径
            sheet_per_table: True 则每个表一个 sheet，False 则全部合并到一个 sheet
        """
        output = Path(output)
        output.parent.mkdir(parents=True, exist_ok=True)

        if not result.tables:
            import openpyxl
            wb = openpyxl.Workbook()
            wb.save(output)
            return output

        import pandas as pd

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            if sheet_per_table:
                for i, df in enumerate(result.tables):
                    sheet_name = f"Table_{i + 1}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._auto_column_width(writer, sheet_name, df)
            else:
                combined = pd.concat(result.tables, ignore_index=True)
                combined.to_excel(writer, sheet_name="All_Tables", index=False)
                self._auto_column_width(writer, "All_Tables", combined)

        return output

    @staticmethod
    def _auto_column_width(writer, sheet_name: str, df) -> None:
        """自动调整列宽 + 数字右对齐。"""
        ws = writer.sheets[sheet_name]
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = len(str(col_name))
            for val in df.iloc[:, col_idx - 1]:
                cell_len = len(str(val)) if val is not None else 0
                max_len = max(max_len, cell_len)
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)
        right_align_numbers(ws)
